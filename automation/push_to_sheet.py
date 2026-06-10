"""Canlı Google Sheet'e gönderici. GCC_CRM.xlsx'teki güncel CRM'i Apps Script web app'e POST eder.
Apps Script v3 TOPLU yazar (clear + tek setValues) -> tek POST'ta tüm liste, binlerce satırda bile hızlı.
Manuel sütunlar (Durum, Not) Sheet tarafında KORUNUR (script öyle yazıldı), o yüzden gönderilmez.
sheet_webapp_url boşsa atlar. build_crm.py sonunda otomatik çağrılır.
"""
import os, json, time, urllib.request
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
CRM = os.path.join(HERE, "..", "GCC_CRM.xlsx")
# (Sheet sütun adı, CRM xlsx sütun adı) — Durum/Not MANUEL, gönderilmez ki Sheet'teki ezilmesin.
AUTO = [("No","No"),("Kategori","Kategori"),("İsim","İsim"),("Rol","Rol"),("Şirket","Şirket"),
        ("Ülke","Ülke"),("Yanıt Olasılığı","Yanıt Olasılığı"),("İş E-postası","İş E-postası"),
        ("LinkedIn","LinkedIn"),("Kabul Durumu","Kabul Durumu"),("Son Temas","Son Temas")]

def main():
    url = json.load(open(os.path.join(HERE,"config.json"), encoding="utf-8")).get("sheet_webapp_url","").strip()
    if not url:
        print("[sheet] sheet_webapp_url bos — canli Sheet'e gonderim atlandi (once Apps Script kur).")
        return
    if not os.path.exists(CRM):
        print("[sheet] GCC_CRM.xlsx yok — once build_crm.py."); return
    ws = load_workbook(CRM)["CRM"]; head = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(head)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[idx["İsim"]]: continue
        rows.append({sheet_h: (r[idx[crm_h]] if r[idx[crm_h]] is not None else "") for sheet_h, crm_h in AUTO})

    # Toplu Apps Script -> TEK POST (parcalama YANLIS olur, her POST sayfayi sifirlayip yeniden yazar).
    body = json.dumps({"rows": rows}).encode("utf-8")
    last = None
    for attempt in range(1, 5):
        req = urllib.request.Request(url, data=body, headers={"Content-Type":"application/json"})
        try:
            with urllib.request.urlopen(req, timeout=180) as resp:
                txt = resp.read().decode()[:200]
                if txt.lstrip().startswith("{"):
                    print(f"[sheet] gonderildi ({len(rows)} satir): {txt}"); return
                print(f"[sheet] deneme {attempt}/4 Apps Script hata sayfasi dondu")
        except Exception as e:
            last = e; print(f"[sheet] deneme {attempt}/4 basarisiz: {e}")
        if attempt < 4: time.sleep(attempt * 20)
    print("[sheet] HATA (4 deneme sonrasi vazgecildi):", last)

if __name__ == "__main__":
    main()

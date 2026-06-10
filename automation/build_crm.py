"""CRM görünümü üretir (yeni sade şema).
Sütunlar: No, Kategori, İsim, Rol, Şirket, Ülke, Yanıt Olasılığı, İş E-postası, LinkedIn,
          Kabul Durumu (bot), Durum (manuel), Son Temas, Not (manuel).
- Kabul Durumu: LinkedIn isteği kabul/gönderim durumu. accepted.json (kabul edenler) > contacted.json (istek
  gönderilenler) > "—". Kabul edenleri LinkedIn izleme botu accepted.json'a yazar.
- Durum / Not: MANUEL — bu dosya yeniden üretilince KORUNUR; canlı Sheet'te de Apps Script korur.
Çıktı: GCC_CRM.xlsx (CRM + Özet). Sonunda canlı Google Sheet'e gönderir.
"""
import csv, os, json, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
_LIVE = os.path.join(HERE, "..", "GCC_TUM_KISILER_LIVE.csv")
_MASTER = os.path.join(HERE, "..", "GCC_TUM_KISILER.csv")
SRC = _LIVE if os.path.exists(_LIVE) else _MASTER
OUT = os.path.join(HERE, "..", "GCC_CRM.xlsx")
LI_DONE = os.path.join(HERE, "contacted.json")      # istek GÖNDERİLENLER
LI_ACCEPTED = os.path.join(HERE, "accepted.json")   # isteği KABUL EDENLER (LinkedIn izleme botu yazar)
FONT = "Arial"; NAVY="1F3864"; GREEN="548235"; ORANGE="C55A11"; WHITE="FFFFFF"
thin = Side(style="thin", color="BFBFBF"); border = Border(left=thin,right=thin,top=thin,bottom=thin)

HEADERS = ["No","Kategori","İsim","Rol","Şirket","Ülke","Yanıt Olasılığı","İş E-postası","LinkedIn",
           "Kabul Durumu","Durum","Son Temas","Not"]
MANUEL = ("Durum","Not")  # kullanicinin elle yazdigi, korunan sutunlar

def load_keys(path):
    if not os.path.exists(path): return {}
    data = json.load(open(path, encoding="utf-8"))
    if isinstance(data, dict): return data
    return {k: "" for k in data}  # liste -> tarihsiz

def preserved():
    """Var olan CRM xlsx'ten manuel sütunları (Durum, Not) anahtar->değer al. Eski 'Genel Durum'u Durum'a tasi."""
    keep = {}
    if os.path.exists(OUT):
        try:
            wb = load_workbook(OUT)
        except Exception:
            return keep
        if "CRM" in wb.sheetnames:
            ws = wb["CRM"]; head = [c.value for c in ws[1]]
            def idx(n): return head.index(n) if n in head else None
            ki, kc = idx("İsim"), idx("Şirket")
            di, gi, ni = idx("Durum"), idx("Genel Durum"), idx("Not")
            for r in ws.iter_rows(min_row=2, values_only=True):
                if ki is None: break
                key = f"{str(r[ki]).strip().lower()}|{str(r[kc]).strip().lower()}"
                durum = (r[di] if di is not None else "") or (r[gi] if gi is not None else "")
                keep[key] = {"durum": durum or "", "not": (r[ni] if ni is not None else "") or ""}
    return keep

def main():
    rows = list(csv.DictReader(open(SRC, encoding="utf-8-sig")))
    lidone, accepted, keep = load_keys(LI_DONE), load_keys(LI_ACCEPTED), preserved()

    wb = Workbook(); ws = wb.active; ws.title = "CRM"; ws.sheet_view.showGridLines = False
    for i,h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name=FONT, bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = border
    ws.freeze_panes = "C2"; ws.row_dimensions[1].height = 30

    n_sent=n_acc=0
    r_ = 2
    for r in rows:
        isim, sirket = r.get("Isim","").strip(), r.get("Sirket","").strip()
        key = f"{isim.lower()}|{sirket.lower()}"
        if key in accepted:
            kabul = "✓ Kabul etti"; n_acc += 1; n_sent += 1
        elif key in lidone:
            kabul = "İstek gönderildi"; n_sent += 1
        else:
            kabul = "—"
        son = accepted.get(key) or lidone.get(key) or ""
        km = keep.get(key, {})
        vals = [r_-1, r.get("Kategori",""), isim, r.get("Rol",""), sirket, r.get("Ulke",""),
                r.get("Yanit Olasiligi",""), r.get("Is E-postasi",""), r.get("LinkedIn",""),
                kabul, km.get("durum",""), son, km.get("not","")]
        for c_,v in enumerate(vals, start=1):
            cell = ws.cell(row=r_, column=c_, value=v); cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="center" if c_ in (1,10) else "left")
            cell.font = Font(name=FONT, size=10)
            if c_==10:  # Kabul Durumu renk
                col = {"✓ Kabul etti":GREEN,"İstek gönderildi":ORANGE}.get(v,"808080")
                cell.font = Font(name=FONT,size=10,bold=str(v).startswith("✓"),color=col)
        ws.row_dimensions[r_].height = 26; r_ += 1
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{r_-1}"
    for i,w in enumerate([4,20,22,20,24,13,15,26,16,16,14,14,28], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Özet
    sm = wb.create_sheet("Özet"); sm.sheet_view.showGridLines=False
    sm.column_dimensions["A"].width=4; sm.column_dimensions["B"].width=36; sm.column_dimensions["C"].width=14
    def line(rr, label, val, color="000000", bold=False):
        sm.cell(row=rr,column=2,value=label).font=Font(name=FONT,size=11,bold=bold,color=color)
        sm.cell(row=rr,column=3,value=val).font=Font(name=FONT,size=11,bold=True,color=color)
    sm.cell(row=2,column=2,value="CRM Özeti").font=Font(name=FONT,bold=True,size=15,color=NAVY)
    line(4,"Toplam kişi", len(rows), NAVY, True)
    line(5,"LinkedIn isteği gönderildi", n_sent, ORANGE)
    line(6,"İsteği kabul eden", n_acc, GREEN)
    line(7,"Güncelleme", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
    sm.cell(row=9,column=2,value="Not: 'Durum' ve 'Not' sütunlarını elle doldur — yeniden üretilince KORUNUR. 'Kabul Durumu' botun işidir.").font=Font(name=FONT,size=9,italic=True,color="808080")

    wb.save(OUT)
    print(f"CRM hazir: {os.path.basename(OUT)} | {len(rows)} kisi, istek gonderilen: {n_sent}, kabul eden: {n_acc}")
    try:
        import push_to_sheet; push_to_sheet.main()  # canli Google Sheet'e gonder
    except Exception as e:
        print(f"[sheet] {e}")

if __name__ == "__main__":
    main()

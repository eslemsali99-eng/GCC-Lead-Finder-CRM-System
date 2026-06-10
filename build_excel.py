import urllib.parse
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from people_data import PEOPLE
from contacts_data import CONTACTS
from executives_data import EXECUTIVES

FONT = "Arial"
NAVY="1F3864"; BLUE="2E5395"; GREY="F2F2F2"; GREEN="548235"; ORANGE="C55A11"; WHITE="FFFFFF"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
link_font = Font(name=FONT, size=10, color="0563C1", underline="single")

def style_header(cell, bg=NAVY):
    cell.font = Font(name=FONT, bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

def cell_set(ws, coord, value, bold=False, size=10, color="000000", bg=None, wrap=True, align="left"):
    c = ws[coord]; c.value = value
    c.font = Font(name=FONT, bold=bold, size=size, color=color)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    return c

def hyperlink(url, text):
    url = url.replace('"', '%22')
    return f'=HYPERLINK("{url}","{text}")'

def li_people(name):
    return "https://www.linkedin.com/search/results/people/?keywords=" + urllib.parse.quote(name)

wb = Workbook()

# ---------- SHEET 1: OKU ----------
ws = wb.active; ws.title = "OKU - Yontem & Yasal"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 112
cell_set(ws, "B2", "GCC Zengin Yatırımcı Lead Sistemi — Önce Bunu Oku", bold=True, size=16, color=NAVY)
rows = [
 ("NE ALDIN", NAVY),
 ("Teknosab arazi + Ege villa + fabrika + Türk vatandaşlığı paketini Körfez'deki zengin kişilere satman için "
  "GERÇEK hedef listesi. 'GCC Hedef Kişiler' sekmesindeki 38 isim ARACI DEĞİL; doğrudan kurucu/başkan/CEO — "
  "yani parayı verecek kişinin kendisi. Hepsi web'den (Forbes, Arabian Business, şirket siteleri) doğrulandı.", None),
 ("", ""),
 ("NEDEN UYDURMA NUMARA YOK — DÜRÜST", ORANGE),
 ("İsim, rol, şirket ve web sitesi gerçek ve public. Ama kişisel telefon/WhatsApp/mail'i uydurmadım — onlar "
  "'Scraping Pipeline' + 'Keşif Yöntemleri' ile public kaynaklardan SEN doldurursun. Sahte numara = yanlış kişiyi "
  "aramak = itibar kaybı. Senin satışın güven üzerine kurulu, onu riske atmam.", None),
 ("", ""),
 ("DOSYAYI VS CODE'DA AÇMA", ORANGE),
 (".xlsx ikili (binary) bir dosyadır; VS Code metin editöründe açarsan anlamsız karakterler görürsün — bozuk "
  "değildir. Dosyaya çift tıkla, Excel veya Numbers ile aç. Hızlı bakış için GCC_Hedef_Kisiler.csv de var.", None),
 ("", ""),
 ("ÖNCE BURADAN BAŞLA", GREEN),
 ("'Yöneticiler (Yüksek Yanıt)' sekmesi = 25 profesyonel CEO/CFO. Patron değiller ama villa/vatandaşlık alacak kadar "
  "yüksek maaşlı VE LinkedIn'de aktif olduğu için GERÇEKTEN cevap verirler. İlk satışları buradan çıkar. "
  "Milyarder sahipler ('GCC Hedef Kişiler') uzun vadeli, düşük yanıtlı hedeftir — sabır ister.", None),
 ("", ""),
 ("NASIL KULLANILIR", NAVY),
 ("1) 'GCC Hedef Kişiler' — 35 milyarder/sahip. Şirket web + LinkedIn + 27'sinde gerçek public mail/telefon dolu.", None),
 ("2) 'Keşif Yöntemleri' — bu kişilerin public olarak NEREDE göründüğünü (liste, etkinlik, borsa, basın...) topla.", None),
 ("3) 'Scraping Pipeline' — Apify ile public iş iletişimini (mail/telefon/sosyal) otomatik doldur.", None),
 ("4) Numarayı 'WhatsApp (ham)' sütununa yapıştır → yanındaki link otomatik tıklanabilir wa.me olur.", None),
 ("5) 'Outreach Stratejisi' — üst düzey kişiye nasıl yazılır (asistan/IR üzerinden, kısa ve değerli).", None),
 ("", ""),
 ("YASAL ÇİZGİ (KVKK/GDPR)", ORANGE),
 ("Public iş iletişimi (kurumsal mail, IR, santral, asistan) toplamak ve B2B teklif yasaldır. İlk mesajda opt-out "
  "ver. Toplu otomatik WhatsApp gönderme (numara banlanır). Her verinin kaynağını 'Kaynak' sütunundan izle.", None),
]
r = 3
for text, color in rows:
    if text == "":
        r += 1; continue
    if color and len(text) < 60:
        cell_set(ws, f"B{r}", text, bold=True, size=12, color=color)
    else:
        cell_set(ws, f"B{r}", text, size=10)
        ws.row_dimensions[r].height = max(15, (len(text)//100 + 1) * 15)
    r += 1

# ---------- SHEET 2: GCC Hedef Kişiler ----------
ws2 = wb.create_sheet("GCC Hedef Kisiler"); ws2.sheet_view.showGridLines = False
headers = ["No","Sektör","İsim","Rol","Şirket","Ülke / Şehir","Zenginlik Sinyali","Yanıt Olasılığı",
           "Şirket Web","LinkedIn (kişi ara)","LinkedIn Şirket","İş E-postası","IR/Basın E-postası","Telefon",
           "WhatsApp (ham)","WhatsApp Linki","Instagram","İletişim Sayfası","Erişim Stratejisi","Öncelik","Doğrulama / Not","Kaynak"]
for i,h in enumerate(headers, start=1):
    style_header(ws2.cell(row=1, column=i, value=h))
ws2.freeze_panes = "C2"; ws2.row_dimensions[1].height = 44
WA_HAM_COL = 15  # O sütunu

listed_kw = ("listel","tadawul","adx","borsa","boursa","qse","dfm","halka")
row = 2
for idx,(sektor,isim,rol,sirket,ulke,zengin,web,durum,kaynak) in enumerate(PEOPLE, start=1):
    listed = any(k in (zengin+durum).lower() for k in listed_kw)
    strat = ("Halka açık: IR maili üzerinden kurumsal giriş + 'Office of the Chairman'/asistan (EA)." if listed
             else "Özel şirket: santral + 'Office of the Chairman'/asistan (EA). Önce LinkedIn'de bağlan.")
    oncelik = "Yüksek" if ("milyar" in zengin.lower() or "forbes" in zengin.lower() or "$" in zengin) else "Orta"
    dom = web.split("//")[-1].replace("www.","").rstrip("/")
    email, ir_email, phone, li_co, ig, contact_pg, note = CONTACTS.get(dom, ("","","","","","",""))
    full_note = (durum + (" | " + note if note else "")).strip(" |")
    ham = get_column_letter(WA_HAM_COL)
    vals = [idx, sektor, isim, rol, sirket, ulke, zengin,
            "Düşük (sahip/başkan — uzun vade)",
            hyperlink(web, dom), hyperlink(li_people(isim), "LinkedIn'de bul"),
            hyperlink(li_co, "Şirket sayfası") if li_co else "",
            email, ir_email, phone, "",
            f'=IF(${ham}{row}="","", HYPERLINK("https://wa.me/"&SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(${ham}{row},"+",""),"-","")," ",""),"WhatsApp\'ta Aç"))',
            ig, hyperlink(contact_pg, "İletişim/IR") if contact_pg else "",
            strat, oncelik, full_note, hyperlink(kaynak, "Kaynak")]
    for c,v in enumerate(vals, start=1):
        ws2.cell(row=row, column=c, value=v)
    row += 1

link_cols = {9,10,11,16,18,22}  # hyperlink içeren sütunlar
for r_ in range(2, row):
    for col in range(1, 23):
        cell = ws2.cell(row=r_, column=col); cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True,
                                   horizontal="center" if col in (1,20) else "left")
        cell.font = link_font if col in link_cols else Font(name=FONT, size=10)
        if col == 20 and cell.value == "Yüksek":
            cell.font = Font(name=FONT, size=10, bold=True, color=GREEN)
    ws2.row_dimensions[r_].height = 60
for i,w in enumerate([4,15,19,18,20,15,30,17,14,14,13,24,22,16,15,14,16,13,30,9,34,8], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ---------- SHEET 2B: Yöneticiler (Yüksek Yanıt) ----------
wx = wb.create_sheet("Yoneticiler (Yuksek Yanit)"); wx.sheet_view.showGridLines = False
cell_set(wx, "A1", "", )
xheaders = ["No","Sektör","İsim","Unvan","Şirket","Ülke / Şehir","Neden Ulaşılabilir","Gelir / Zenginlik Sinyali",
            "LinkedIn (profil)","Şirket Web","İş E-postası (scrape)","WhatsApp (ham)","WhatsApp Linki",
            "Erişim Stratejisi","Durum / Not","Kaynak"]
# Üstte açıklama satırı
wx.merge_cells("A1:P1")
cell_set(wx, "A1", "İLK TEMAS ÖNCELİĞİ: Bu kişiler profesyonel, yüksek maaşlı ve LinkedIn'de aktif — patronlardan ÇOK daha hızlı cevap verirler. Önce buradan başla. (Sahip değiller; gayrimenkul şirketleri çıkarıldı.)",
         bold=True, size=11, color=WHITE, bg=GREEN, align="left")
wx.row_dimensions[1].height = 30
for i,h in enumerate(xheaders, start=1):
    style_header(wx.cell(row=2, column=i, value=h), bg=GREEN)
wx.freeze_panes = "C3"; wx.row_dimensions[2].height = 42
XWA = 12  # WhatsApp ham sütunu (L)
xrow = 3
for idx,(sektor,isim,unvan,sirket,ulke,neden,gelir,linkedin,web,durum,kaynak) in enumerate(EXECUTIVES, start=1):
    dom = web.split("//")[-1].replace("www.","").rstrip("/")
    ham = get_column_letter(XWA)
    li_text = "LinkedIn profili" if "/in/" in linkedin else "LinkedIn'de bul"
    vals = [idx, sektor, isim, unvan, sirket, ulke, neden, gelir,
            hyperlink(linkedin, li_text), hyperlink(web, dom), "", "",
            f'=IF(${ham}{xrow}="","", HYPERLINK("https://wa.me/"&SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(${ham}{xrow},"+",""),"-","")," ",""),"WhatsApp\'ta Aç"))',
            "LinkedIn'de bağlan + kısa not; iş maili için şirket deseni/scrape. Profesyonel, randevuya açık.",
            durum, hyperlink(kaynak, "Kaynak")]
    for c,v in enumerate(vals, start=1):
        wx.cell(row=xrow, column=c, value=v)
    xrow += 1
xlink_cols = {9,10,13,16}
for r_ in range(3, xrow):
    for col in range(1, 17):
        cell = wx.cell(row=r_, column=col); cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="center" if col==1 else "left")
        cell.font = link_font if col in xlink_cols else Font(name=FONT, size=10)
    wx.row_dimensions[r_].height = 52
for i,w in enumerate([4,15,20,17,20,16,28,30,15,15,18,15,14,28,30,8], start=1):
    wx.column_dimensions[get_column_letter(i)].width = w

# ---------- SHEET 2C: Borsa Kurul & C-Suite (Geniş Liste) ----------
BIGLIST = json.load(open("biglist.json", encoding="utf-8"))
ULKE_TR = {"Saudi Arabia":"Suudi Arabistan","UAE":"BAE","Qatar":"Katar","Kuwait":"Kuveyt"}
ROYAL = ("sheikh","prince","hrh","hh ","sheikha")
def responsiveness(role, name):
    r = role.lower(); n = name.lower()
    if any(k in r for k in ("ceo","cfo","managing director"," md","chief","president","general manager","group ce")):
        return "Yüksek (profesyonel)"
    if "chair" in r:
        return "Düşük (royal/sahip)" if any(k in n for k in ROYAL) else "Orta (başkan)"
    return "Orta"

wb_big = wb.create_sheet("Borsa Kurul & C-Suite"); wb_big.sheet_view.showGridLines = False
wb_big.merge_cells("A1:K1")
cell_set(wb_big, "A1", f"GENİŞ LİSTE — {len(BIGLIST)} kişi: GCC borsalarındaki büyük şirketlerin Başkan/CEO/CFO'ları (public kurul & IR sayfaları). "
                       "CEO/CFO = yüksek yanıt; royal başkanlar = düşük yanıt. Toplu lead; ilk mesaj öncesi unvanı 'Kaynak'tan teyit et.",
         bold=True, size=11, color=WHITE, bg=NAVY, align="left")
wb_big.row_dimensions[1].height = 30
bheaders = ["No","Sektör","İsim","Rol","Şirket","Ülke","Yanıt Olasılığı","LinkedIn","Şirket Web","İş E-postası (scrape)","Kaynak"]
for i,h in enumerate(bheaders, start=1):
    style_header(wb_big.cell(row=2, column=i, value=h), bg=NAVY)
wb_big.freeze_panes = "C3"; wb_big.row_dimensions[2].height = 30
brow = 3
for idx,p in enumerate(BIGLIST, start=1):
    li = p.get("linkedin_url",""); web = p.get("company_website",""); src = p.get("source","")
    dom = web.split("//")[-1].replace("www.","").rstrip("/") if web else ""
    li_text = "LinkedIn profili" if "/in/" in li else "LinkedIn'de bul"
    vals = [idx, p.get("sector",""), p["full_name"], p.get("role",""), p["company"],
            ULKE_TR.get(p.get("country",""), p.get("country","")), responsiveness(p.get("role",""), p["full_name"]),
            hyperlink(li, li_text) if li else "", hyperlink(web, dom) if web else "", "",
            hyperlink(src, "Kaynak") if src else ""]
    for c,v in enumerate(vals, start=1):
        wb_big.cell(row=brow, column=c, value=v)
    brow += 1
blink = {8,9,11}
for r_ in range(3, brow):
    yr = wb_big.cell(r_,7).value
    for col in range(1,12):
        cell = wb_big.cell(row=r_, column=col); cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="center" if col in (1,7) else "left")
        cell.font = link_font if col in blink else Font(name=FONT, size=10)
        if col==7 and yr and yr.startswith("Yüksek"):
            cell.font = Font(name=FONT, size=10, bold=True, color=GREEN)
    wb_big.row_dimensions[r_].height = 30
for i,w in enumerate([4,20,24,22,26,15,20,15,18,20,9], start=1):
    wb_big.column_dimensions[get_column_letter(i)].width = w

# ---------- SHEET 3: Keşif Yöntemleri ----------
wk = wb.create_sheet("Kesif Yontemleri"); wk.sheet_view.showGridLines = False
for col,w in zip("ABCDE",[3,26,40,40,10]): wk.column_dimensions[col].width = w
cell_set(wk, "B2", "Keşif Yöntemleri — Kurucu/CEO Public Olarak Nerede Görünür?", bold=True, size=14, color=NAVY)
cell_set(wk, "B3", "Mantık: Zengin Arap iş insanı bulunduğu yeri/işini büyük oranda PUBLIC paylaşır. Bu sinyalleri toplar, "
                   "kişiyi tespit eder, public iş iletişimine ulaşırsın. Aşağıdaki her yöntem 'Scraping Pipeline' ile beslenir.",
         size=10, color=ORANGE)
kh = ["No","Yöntem","Neden Public Paylaşılır","Nasıl Veri Çekilir","Öncelik"]
for i,h in enumerate(kh, start=1):
    style_header(wk.cell(row=5, column=i+1, value=h), bg=BLUE)
wk.row_dimensions[5].height = 34
methods = [
 ("Zenginler Listeleri & Ödüller","Forbes Middle East, Arabian Business, Gulf Business '100 Most Powerful Arabs', 'Top CEO Awards' her yıl isim+şirketi yayınlar — prestij olduğu için kişiler paylaşır.","Liste/ödül sayfalarını Apify website-content-crawler ile tara; isim+şirket çıkar → bu listeye ekle.","Yüksek"),
 ("Etkinlik & Fuar Katılımı","FII, AIM Congress, Cityscape, LEAP'te konuşmacı/sergici/'attending' olmak görünürlük demek; LinkedIn Events'te katılım public.","Konuşmacı & sergici listeleri + LinkedIn etkinlik attendees ('Etkinlik & Sinyal' sekmesi).","Yüksek"),
 ("Borsa & Yatırımcı İlişkileri (IR)","Tadawul, ADX, DFM, Boursa Kuwait, QSE şirketleri yönetim kurulu üyelerini + hisse oranlarını RESMEN açıklar; IR maili public.","Borsa/şirket 'Board of Directors' & 'Investor Relations' sayfalarını scrape → isim + resmi IR iletişimi.","Yüksek"),
 ("Şirket Basın Bültenleri & Yatırım Haberleri","Kurucu yeni yatırım/genişleme/satın alma açıkladığında PR ve haberde adı geçer — likidite ve niyet sinyali.","Google News + Apify news scraper; '<sektör> investment GCC', 'acquisition UAE' sorguları; alert kur.","Yüksek"),
 ("Ticaret Odası & İş Konseyi Dizinleri","Qatari Businessmen Association, Suudi/Dubai Ticaret Odaları, US-Saudi/Türk-Arap iş konseyleri üye dizinleri public.","Üye dizini sayfalarını scrape; iş konseyi etkinlik katılımcı listeleri.","Yüksek"),
 ("Vakıf & Hayırseverlik Kurulları","Community Jameel, Alghanim/aile vakıfları yönetim kurullarında zengin isimler yer alır; itibar için açıklanır.","Vakıf 'Board/Trustees' sayfalarını scrape → isim → şirket → public iletişim.","Orta"),
 ("Aile Şirketi Ağları","Tharawat, Pearl Initiative, Hawkamah, family office forumları üye/konuşmacı listeleri — aile serveti sinyali.","Ağ/forum üye & konuşmacı sayfalarını scrape.","Orta"),
 ("Lüks Varlık Sinyalleri","Dubai World Cup at sahipleri, yat/jet, sanat müzayedesi (Sotheby's/Christie's ME) alıcıları, polo kulüpleri — servet göstergesi ve çoğu public.","Etkinlik kazanan/katılımcı sayfaları, müzayede sonuç listeleri, kulüp üye haberleri.","Orta"),
 ("Sosyal Medya Bio & Gönderiler","Doğrulanmış Instagram/X iş hesapları; bio'da iş maili/WhatsApp; gönderilerde seyahat/yatırım/proje sinyali.","Apify Instagram/X scraper ile bio + iletişim linki + konum/etkinlik etiketleri.","Orta"),
 ("Podcast & Röportajlar","Kurucular röportaj/podcast'te plan, ilgi alanı, bazen iletişim paylaşır — sıcaklık ve açı verir.","YouTube/podcast açıklamaları + transkript; isim+bağlam çıkar, kişiselleştirme için kullan.","Orta"),
 ("Mega-Proje Konsorsiyumları & İhaleler","NEOM, Expo legacy, dev altyapı projelerinde ortak firmalar ve sahipleri resmen açıklanır.","Proje/ihale duyuru sayfaları + haber; ortak firma → sahip/başkan.","Orta"),
 ("Sponsorluk & İsim Hakları","Bina/stadyum/üniversite isim hakları zengin aileleri işaret eder; basında geçer.","Haber + kurum sayfaları; isim → aile/şirket → public iletişim.","Düşük"),
]
rr = 6
for i,(name,why,how,pri) in enumerate(methods, start=1):
    wk.cell(row=rr, column=2, value=i)
    wk.cell(row=rr, column=3, value=name)
    wk.cell(row=rr, column=4, value=why)
    wk.cell(row=rr, column=5, value=how)
    wk.cell(row=rr, column=6, value=pri)
    for col in range(2,7):
        cell = wk.cell(row=rr, column=col); cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="center" if col in (2,6) else "left")
        cell.font = Font(name=FONT, size=10, bold=(col==3))
        if col==6 and pri=="Yüksek": cell.font = Font(name=FONT, size=10, bold=True, color=GREEN)
    wk.row_dimensions[rr].height = 58
    rr += 1

# ---------- SHEET 4: Etkinlik & Sinyal Hedefleme ----------
wse = wb.create_sheet("Etkinlik & Sinyal"); wse.sheet_view.showGridLines = False
for col,w in zip("ABCDEFG",[3,30,15,13,32,16,34]): wse.column_dimensions[col].width = w
cell_set(wse, "B2", "Etkinlik & Sinyal Hedefleme — CEO'ları Bulundukları Yerden Yakala", bold=True, size=14, color=NAVY)
cell_set(wse, "B3", "Kişi bir etkinliğe katıldığını PUBLIC paylaşır (LinkedIn 'Attending', konuşmacı/sergici listesi). Bu sinyali "
                    "topla → kişisel LinkedIn profili → public iletişim.", size=10, color=ORANGE)
ehead = ["No","Etkinlik","Şehir","Dönem","Neden Senin Alıcın Burada","Web","İsim Nasıl Çıkarılır"]
for i,h in enumerate(ehead, start=1):
    style_header(wse.cell(row=5, column=i+1, value=h), bg=BLUE)
wse.row_dimensions[5].height = 38
events = [
 ("Future Investment Initiative (FII)","Riyad","Ekim","'Çölün Davos'u' — sovereign fonlar, family office'ler, ultra-zenginler","fii-institute.org","Konuşmacı+partner listesi public; LinkedIn'de 'attending'"),
 ("AIM Congress","Abu Dabi","Mayıs","FDI/yatırım odaklı; ülke-proje arayan sermaye = arazi/fabrika alıcısı","aimcongress.com","Sergici+delege dizini; LinkedIn etkinlik"),
 ("Cityscape Global","Dubai","Kasım","Ortadoğu'nun en büyük gayrimenkul fuarı; yatırımcı yoğun","cityscapeglobal.com","Sergici listesi; ziyaretçi randevu"),
 ("World Government Summit","Dubai","Şubat","Üst düzey iş+devlet ağı; ultra-HNW karar vericiler","worldgovernmentsummit.org","Konuşmacı + LinkedIn attendees"),
 ("LEAP","Riyad","Şubat","Dev teknoloji & yatırım zirvesi; genç zengin yatırımcı havuzu","onegiantleap.com","Sergici+speaker; LinkedIn etkinlik"),
 ("Qatar Economic Forum","Doha","Mayıs","Bloomberg destekli; bölgesel sermaye sahipleri ve CEO'lar","qatareconomicforum.com","Konuşmacı listesi → LinkedIn profil"),
 ("Sharjah FDI Forum","Şarjah","Değişken","Sanayi/fabrika yatırımı arayan sermaye = fabrika paketine birebir","investinsharjah.ae","Katılımcı+partner listesi"),
 ("LinkedIn Events (arama)","Online","Sürekli","'Turkey real estate investment', 'family office UAE' aramaları","linkedin.com/search/results/events","'Attending' diyenler PUBLIC → kişisel profil → iletişim"),
]
for i,(name,city,period,why,web,how) in enumerate(events, start=1):
    rr = 5+i
    for col,val in zip(range(2,9),[i,name,city,period,why,hyperlink(f"https://{web}",web),how]):
        cell = wse.cell(row=rr, column=col, value=val); cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="center" if col in (2,4,5) else "left")
        cell.font = link_font if col==7 else Font(name=FONT, size=10)
    wse.row_dimensions[rr].height = 42
base = 5+len(events)+2
cell_set(wse, f"B{base}", "ZENGİNLİK SİNYALLERİ — Bir kişiyi 'hedef' yapan parametreler", bold=True, size=11, color=WHITE, bg=BLUE)
wse.merge_cells(f"B{base}:H{base}")
signals = [
 ("O bölgede şirketi/holdingi olan","GCC serbest bölge / DIFC / ADGM kayıtlı = likidite yüksek"),
 ("Etkinlikte konuşmacı/sergici","Public görünürlük + bütçe = ciddi oyuncu"),
 ("Forbes/Arabian Business listesinde","Doğrulanmış servet sinyali"),
 ("Family office / Chairman / Founder unvanı","LinkedIn unvanı doğrudan karar verici"),
 ("Mevcut yurtdışı gayrimenkulü/yatırımı","Zaten sınır ötesi yatırım alışkanlığı var"),
 ("Vatandaşlık/golden visa ilgisi","İlgili içerik/etkinlik = sıcak lead"),
]
for j,(s,d) in enumerate(signals):
    rr = base+1+j
    cell_set(wse, f"B{rr}", "✓ "+s, bold=True, size=10, color=NAVY, bg=GREY); wse.merge_cells(f"B{rr}:C{rr}")
    cell_set(wse, f"D{rr}", d, size=10); wse.merge_cells(f"D{rr}:H{rr}")
    wse.row_dimensions[rr].height = 24

# ---------- SHEET 5: Scraping Pipeline ----------
ws3 = wb.create_sheet("Scraping Pipeline"); ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width=3; ws3.column_dimensions["B"].width=30; ws3.column_dimensions["C"].width=82
cell_set(ws3, "B2", "Apify Scraping Pipeline — Public İletişimi Otomatik Doldur", bold=True, size=14, color=NAVY)
cell_set(ws3, "B3", "Script: apify_pipeline.js. APIFY_TOKEN koy, node ile çalıştır. Sadece public iş verisi. Önce MAX_ITEMS=10 ile test et.", size=10, color=ORANGE)
pipe = [
 ("Komut / Aktör","Ne yapar — hangi sütunu doldurur", True),
 ("enrich\n(contact-info-scraper)","'GCC Hedef Kişiler'deki şirket sitelerini gezer; kurumsal e-posta, telefon, sosyal medya → İş E-postası/Telefon/Instagram.", False),
 ("maps\n(crawler-google-places)","'family office Dubai', 'holding company Riyadh' sorgularıyla işletme adı/telefon/website/WhatsApp → yeni hedef.", False),
 ("events\n(website-content-crawler)","FII/AIM/Cityscape konuşmacı+sergici sayfalarından isim çıkar → LinkedIn kişisel profile bağla → enrich.", False),
 ("richlist\n(website-content-crawler)","Forbes ME / Arabian Business / Gulf Business liste & ödül sayfalarını tarar; isim+şirket çıkarır → listeye ekle.", False),
 ("news\n(website-content-crawler)","'<sektör> investment GCC', 'acquisition UAE' haber sorguları; yatırım açıklayan kurucuları yakalar.", False),
 ("social\n(instagram-scraper)","Marka/kişi Instagram bio'sundan iş maili/WhatsApp linki + konum/etkinlik etiketleri.", False),
 ("", "", None),
 ("AKIŞ","enrich ile mevcut 38 kişiyi zenginleştir → richlist/maps/news/events ile yeni hedef üret → 'Keşif Yöntemleri'ndeki diğer kaynakları manuel besle.", True),
 ("MALİYET KONTROL","Her komut MAX_ITEMS ile sınırlı. Apify panelinde aylık spending limit koy. Cron KURMA, manuel tetikle. Toplu çalıştırma öncesi maliyet tahmini al.", True),
 ("YASAL","Sadece public iş verisi. Pazarlamadan önce opt-out'lu ilk mesaj. Toplu otomatik WhatsApp YOK (ban).", True),
]
rr = 5
for c1,c2,hdr in pipe:
    if c1=="" and c2=="": rr+=1; continue
    if hdr is True:
        cell_set(ws3, f"B{rr}", c1, bold=True, size=10, color=WHITE, bg=BLUE)
        cell_set(ws3, f"C{rr}", c2, bold=True, size=10, color=WHITE, bg=BLUE)
    else:
        cell_set(ws3, f"B{rr}", c1, bold=True, size=10, color=NAVY, bg=GREY)
        cell_set(ws3, f"C{rr}", c2, size=10)
    ws3.row_dimensions[rr].height = max(30, (len(c2)//80 + 1) * 16)
    rr += 1

# ---------- SHEET 6: Outreach Stratejisi ----------
ws4 = wb.create_sheet("Outreach Stratejisi"); ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width=3; ws4.column_dimensions["B"].width=24; ws4.column_dimensions["C"].width=92
cell_set(ws4, "B2", "Outreach Stratejisi — Üst Düzey Kişiye Nasıl Ulaşılır", bold=True, size=14, color=NAVY)
out = [
 ("Durum","Yaklaşım", True),
 ("Halka açık şirket başkanı","Yatırımcı İlişkileri (IR) public maili üzerinden kurumsal, kısa giriş; 'yönetim kurulu başkanının ofisine iletilmesi ricasıyla' notu. 1 sayfalık yatırım özeti ekle."),
 ("Özel/aile şirketi sahibi","'Office of the Chairman' / yönetici asistanı (EA) köprüdür. Önce LinkedIn'de bağlan, ortak bağlantı/etkinlik üzerinden gir."),
 ("LinkedIn ilk mesaj","Satış değil, fırsat dili: 'Türkiye'de getiri + vatandaşlık opsiyonlu seçili bir gayrimenkul portföyü; 1 sayfalık özet göndereyim mi?' Kişiye özel 1 cümle (son yatırımı/haberi)."),
 ("Etkinlikte (FII/Cityscape)","Yüz yüze 60 saniyelik pitch + QR ile paket. Önce randevu/networking listesinden iste."),
 ("", ""),
 ("İLK MESAJ KURALLARI","", True),
 ("Kanal","Önce e-posta/LinkedIn (kurumsal güven). WhatsApp yalnızca public iş hattı veya karşı taraf paylaştıysa."),
 ("Kişiselleştir","Her mesajda kuruma/kişiye özel 1 cümle. Kopyala-yapıştır spam algılanır."),
 ("Opt-out","'İlgilenmiyorsanız tek kelime yazın, bir daha rahatsız etmem.' KVKK/GDPR + itibar."),
 ("Takip","3 dokunuş: mesaj → 3 gün → değer (örnek paket) → 5 gün → son hatırlatma. Sonra bırak."),
]
rr = 4
for c1,c2,*hdr in out:
    is_hdr = bool(hdr) and hdr[0] is True
    if c1=="" and c2=="": rr+=1; continue
    if is_hdr:
        cell_set(ws4, f"B{rr}", c1, bold=True, size=10, color=WHITE, bg=BLUE)
        cell_set(ws4, f"C{rr}", c2, bold=True, size=10, color=WHITE, bg=BLUE)
    else:
        cell_set(ws4, f"B{rr}", c1, bold=True, size=10, color=NAVY, bg=GREY)
        cell_set(ws4, f"C{rr}", c2, size=10)
    ws4.row_dimensions[rr].height = max(28, (len(str(c2))//90 + 1) * 16)
    rr += 1

# ---------- MASTER: TÜM KİŞİLER (Kategorili, tek liste) ----------
master = wb.create_sheet("TUM KISILER", index=1); master.sheet_view.showGridLines = False
master.merge_cells("A1:O1")
cell_set(master, "A1", "TÜM KİŞİLER — tek liste, kategori sütunlu. Filtre/sıralama için B (Kategori) ve H (Yanıt Olasılığı) sütununu kullan. "
                       "Detaylı alt-görünümler ayrı sekmelerde.", bold=True, size=11, color=WHITE, bg=NAVY, align="left")
master.row_dimensions[1].height = 26
mheaders = ["No","Kategori","Sektör","İsim","Rol / Unvan","Şirket","Ülke","Yanıt Olasılığı","LinkedIn","Şirket Web","İş E-postası","Telefon","Instagram","Doğrulama / Not","Kaynak"]
for i,h in enumerate(mheaders, start=1):
    style_header(master.cell(row=2, column=i, value=h), bg=NAVY)
master.freeze_panes = "D3"; master.row_dimensions[2].height = 30

rows_master = []
seen_master = set()
def add_master(kategori, sektor, isim, rol, sirket, ulke, yanit, linkedin, web, email, phone, ig, note, kaynak):
    key = (isim.strip().lower(), sirket.strip().lower())
    if key in seen_master: return
    seen_master.add(key)
    rows_master.append((kategori, sektor, isim, rol, sirket, ulke, yanit, linkedin, web, email, phone, ig, note, kaynak))

# 1) Milyarder sahip/kurucu
for (sektor,isim,rol,sirket,ulke,zengin,web,durum,kaynak) in PEOPLE:
    dom = web.split("//")[-1].replace("www.","").rstrip("/")
    email, ir_email, phone, li_co, ig, contact_pg, note = CONTACTS.get(dom, ("","","","","","",""))
    add_master("Milyarder Sahip/Kurucu", sektor, isim, rol, sirket, ulke, "Düşük (sahip — uzun vade)",
               li_people(isim), web, email or ir_email, phone, ig, (durum+(" | "+note if note else "")).strip(" |"), kaynak)
# 2) Profesyonel yönetici (curated, yüksek yanıt)
for (sektor,isim,unvan,sirket,ulke,neden,gelir,linkedin,web,durum,kaynak) in EXECUTIVES:
    add_master("Profesyonel Yönetici", sektor, isim, unvan, sirket, ulke, "Yüksek (profesyonel)",
               linkedin, web, "", "", "", durum, kaynak)
# 3) Borsa kurul & C-suite
for p in BIGLIST:
    add_master("Borsa Kurul/C-Suite", p.get("sector",""), p["full_name"], p.get("role",""), p["company"],
               ULKE_TR.get(p.get("country",""), p.get("country","")), responsiveness(p.get("role",""), p["full_name"]),
               p.get("linkedin_url",""), p.get("company_website",""), "", "", "", "", p.get("source",""))

cat_fill = {"Milyarder Sahip/Kurucu":"FCE4D6","Profesyonel Yönetici":"E2EFDA","Borsa Kurul/C-Suite":"DDEBF7"}
mrow = 3
for (kategori, sektor, isim, rol, sirket, ulke, yanit, linkedin, web, email, phone, ig, note, kaynak) in rows_master:
    dom = web.split("//")[-1].replace("www.","").rstrip("/") if web else ""
    li_text = "LinkedIn profili" if "/in/" in linkedin else "LinkedIn'de bul"
    vals = [mrow-2, kategori, sektor, isim, rol, sirket, ulke, yanit,
            hyperlink(linkedin, li_text) if linkedin else "", hyperlink(web, dom) if web else "",
            email, phone, ig, note, hyperlink(kaynak, "Kaynak") if kaynak else ""]
    for c,v in enumerate(vals, start=1):
        master.cell(row=mrow, column=c, value=v)
    mrow += 1
mlink = {9,10,15}
for r_ in range(3, mrow):
    kat = master.cell(r_,2).value; yan = master.cell(r_,8).value
    for col in range(1,16):
        cell = master.cell(row=r_, column=col); cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="center" if col in (1,) else "left")
        cell.font = link_font if col in mlink else Font(name=FONT, size=10)
        if col==2: cell.fill = PatternFill("solid", fgColor=cat_fill.get(kat,"FFFFFF"))
        if col==8 and yan and yan.startswith("Yüksek"): cell.font = Font(name=FONT, size=10, bold=True, color=GREEN)
    master.row_dimensions[r_].height = 30
master.auto_filter.ref = f"A2:O{mrow-1}"
for i,w in enumerate([4,20,18,22,22,24,13,20,15,17,22,16,15,30,8], start=1):
    master.column_dimensions[get_column_letter(i)].width = w
MASTER_ROWS = rows_master

wb.save("GCC_Yatirimci_Lead_Sistemi.xlsx")

# CSV (hızlı bakış için)
import csv
with open("GCC_Hedef_Kisiler.csv","w",newline="",encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["No","Sektor","Isim","Rol","Sirket","Ulke","Zenginlik Sinyali","Is E-postasi","IR/Basin E-postasi","Telefon","LinkedIn Sirket","Instagram","Iletisim Sayfasi","LinkedIn (kisi ara)","Dogrulama Notu","Kaynak"])
    for idx,(sektor,isim,rol,sirket,ulke,zengin,web,durum,kaynak) in enumerate(PEOPLE, start=1):
        dom = web.split("//")[-1].replace("www.","").rstrip("/")
        email, ir_email, phone, li_co, ig, contact_pg, note = CONTACTS.get(dom, ("","","","","","",""))
        full_note = (durum + (" | " + note if note else "")).strip(" |")
        w.writerow([idx,sektor,isim,rol,sirket,ulke,zengin,email,ir_email,phone,li_co,ig,contact_pg,li_people(isim),full_note,kaynak])

with open("GCC_Yoneticiler.csv","w",newline="",encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["No","Sektor","Isim","Unvan","Sirket","Ulke","Neden Ulasilabilir","Gelir Sinyali","LinkedIn","Sirket Web","Durum","Kaynak"])
    for idx,(sektor,isim,unvan,sirket,ulke,neden,gelir,linkedin,web,durum,kaynak) in enumerate(EXECUTIVES, start=1):
        w.writerow([idx,sektor,isim,unvan,sirket,ulke,neden,gelir,linkedin,web,durum,kaynak])

with open("GCC_Borsa_Kurul_C-Suite.csv","w",newline="",encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["No","Sektor","Isim","Rol","Sirket","Ulke","Yanit Olasiligi","LinkedIn","Sirket Web","Kaynak"])
    for idx,p in enumerate(BIGLIST, start=1):
        w.writerow([idx,p.get("sector",""),p["full_name"],p.get("role",""),p["company"],
                    ULKE_TR.get(p.get("country",""),p.get("country","")),responsiveness(p.get("role",""),p["full_name"]),
                    p.get("linkedin_url",""),p.get("company_website",""),p.get("source","")])

with open("GCC_TUM_KISILER.csv","w",newline="",encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["No","Kategori","Sektor","Isim","Rol","Sirket","Ulke","Yanit Olasiligi","LinkedIn","Sirket Web","Is E-postasi","Telefon","Instagram","Not","Kaynak"])
    for i,(kategori,sektor,isim,rol,sirket,ulke,yanit,linkedin,web,email,phone,ig,note,kaynak) in enumerate(MASTER_ROWS, start=1):
        w.writerow([i,kategori,sektor,isim,rol,sirket,ulke,yanit,linkedin,web,email,phone,ig,note,kaynak])

print(f"saved — MASTER {len(MASTER_ROWS)} tekil kisi ({len(PEOPLE)} sahip + {len(EXECUTIVES)} yonetici + {len(BIGLIST)} borsa, tekrarsiz)")
